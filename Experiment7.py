import math
from itertools import combinations
from itertools import permutations

import openpyxl

wb=openpyxl.load_workbook("Experiment6.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column


def read_data_in_dict():
  trans = []
  for i in range(2,row+1):
    x = sheet_obj.cell(row=i,column=1)
    trans.append(x.value)

  string_data=[]
  for i in range(2,row+1):
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        string_data.append(x.value)
  items = set(string_data)
  items = sorted(items)

  length = len(items)
  transactions=[]
  for i in range(2,row+1):
    s=[]
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        s.append(x.value)
    s=sorted(s)  
    req=[]
    k=0
    for j in range(0,length):
      if k<len(s):
        if items[j]==s[k]:
          req.append(1)
          k=k+1
        else:
          req.append(0)
      else:
        req.append(0)

    transactions.append(req)

  data={
      'items':items,
      'transactions': transactions
  }
  return data
  
data = read_data_in_dict()

def frequent_itemsets(data,level,min_support):
    items = data['items']
    transactions = data['transactions']
    min_freq = math.ceil(min_support*len(transactions))
    sets = list(combinations(items,level))
    frequent_sets = []
    for s in sets:
        freq=get_freq(s,items,transactions)
        if freq>=min_freq:
            frequent_sets.append(s)
    return frequent_sets
        


def get_freq(s,items,transactions):
    freq=0
    for t in transactions:
        temp=1
        for item in s:
            temp*=t[items.index(item)]
        if temp==1:
            freq+=1  
    return freq


def generate_rule(data,min_support,min_confidence):
    itr=row+2
    items = data['items']
    bvb=len(items)
    transactions = data['transactions']
    for l in range(2,bvb+1):
        frequent_set = frequent_itemsets(data,l,min_support)
        for s in frequent_set:
            freq_s = get_freq(s,items,transactions)
            _s_permute = list(permutations(s))
            print(_s_permute)
            for _s in _s_permute:
                for i in range(0,len(_s)-1): 
                    x = _s[0:i+1]
                    y = _s[i+1:]
                    freq_x = get_freq(x,items,transactions)
                    c = freq_s/freq_x
                    if c >= min_confidence:
                      val0 = sheet_obj.cell(row=itr,column=1)
                      val1 = sheet_obj.cell(row=itr,column=2)
                      itr=itr+1
                      val0.value = str(x)+"->"+str(y)
                      val1.value = "confidence is:"+str(c)
                      print(x,' -> ',y,'confidence is ',c)



min_sup=float(input("Enter Min Support: "))
min_conf=float(input("Enter Min Confidence: "))
print("Min Support: ",min_sup)
print("Min Confidence: ",min_conf)

generate_rule(data,min_sup,min_conf)
wb.save("Experiment7_Ans.xlsx")
