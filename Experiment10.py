import openpyxl

wb=openpyxl.load_workbook("/content/Exp10.xlsx") 
sheet = wb.active
row = sheet.max_row
col = sheet.max_column

# blank=openpyxl.load_workbook("/content/blank.xlsx")

def get_min(a,b):
  if a<b:
    return a
  else:
    return b

while row>2:
  matrix=[]
  min=9999999
  x=0
  y=0

  rT=[]
  for i in range(2,col+1):
    rT.append(sheet.cell(row=1,column=i).value)

  for i in range(2,row+1):
    arr=[]
    for j in range(2,col+1):
      var= sheet.cell(row=i,column=j).value
      if var or var==0:
        var=float(var)
        arr.append(var)
        if var<min and var!=0:
          min=var
          x=j-1
          y=i-1
    matrix.append(arr)
  
  n=len(matrix)
  
  mn=0
  exc=0

  if x<y:
    mn=x-1
    exc=y-1
  else:
    mn=y-1
    exc=x-1


  answer=[]
  for i in range(0,n):
    arr=[]
    for j in range(0, len(matrix[i])):
      if i==exc:
        break
      elif i==mn:
        if (i>exc and j!=exc) or i<exc:
          if len(matrix[exc])>j:
            arr.append(get_min( matrix[i][j], matrix[exc][j] ))
          else:
            arr.append(get_min( matrix[i][j], matrix[j][exc] ))
      elif j==mn:
        if (i>exc and j!=exc) or i<exc:
          if len(matrix[exc])>i:
            arr.append(get_min( matrix[i][j], matrix[exc][i] ))
          else:
            arr.append(get_min( matrix[i][j], matrix[i][exc] ))
      elif j==exc and i>exc:
        continue
      else:
        arr.append(matrix[i][j])
    if len(arr)>=1:
      answer.append(arr)
  # print(answer)
  

  try:
    rT[mn]="("+rT[mn]+","+rT[exc]+")"
  except:
    break
  print(rT[mn])
  rT.pop(exc)

  sheet.delete_rows(1,sheet.max_row+1)
  wb.save("/content/Exp10.xlsx")
  row = sheet.max_row
  print(row)

  l=len(rT)
  for i in range(0,l):
    sheet.cell(row=1,column=i+2).value = rT[i]
    sheet.cell(row=i+2,column=1).value = rT[i]
  # print(rT)
  rT=[]
  for i in range(0,len(answer)):
    for j in range(0,len(answer[i])):
      sheet.cell(row=i+2,column=j+2).value = answer[i][j]
  wb.save("/content/Exp10.xlsx")
  wb.save("/content/"+str(cnt)+".xlsx")
  cnt=cnt+1
  row = sheet.max_row
  print(row)