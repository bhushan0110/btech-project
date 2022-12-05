import openpyxl
import math
wb=openpyxl.load_workbook("gini_numerical.xlsx")
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

def getUnique(i):
  S=[]
  for j in range(2,row):
    S.append(sheet.cell(row=j,column=i).value)
  s=set(S)
  return s

x=input("Enter column no:")
y=input("Enter Class column no:")
x=int(x)
y=int(y)

Class = []
Class = getUnique(y)
Class=list(Class)

tmp = sheet.cell(row=2,column=x).value

Info_total=0
a1=0
a2=0
for i in range(2,row+1):
  j = sheet.cell(row=i,column=y).value
  if Class[0]==j:
    a1+=1
  else:
    a2+=1
n=row-1
Info_total= -((a1/n)*math.log2(a1/n) + (a2/n)*math.log2(a2/n))

if type(tmp)==int:
  # CODE FOR NUMERICAL 
  attr = input("Enter attribiute value:")
  attr = int(attr)

  g1=0
  g2=0
  l1=0
  l2=0

  for i in range(2,row+1):
    val=sheet.cell(row=i,column=x).value
    if(val>attr):
      if(sheet.cell(row=i,column=y).value==Class[0]):
          g1+=1
      else:
          g2+=1

    else:
      
      if(sheet.cell(row=i,column=y).value==Class[0]):
          l1+=1
      else:
          l2+=1
  print("Value greater than attribute value and "+Class[0]+" ",g1)
  print("Value greater than attribute value and "+Class[1]+" ",g2)
  print("Value lesser than attribute value and "+Class[0]+" ",l1)
  print("Value lesser than attribute value and "+Class[1]+" ",l2)

  Info_gain=0
  gini=0
  gSum=(g1+g2)/n
  lSum=(l1+l2)/n

  Info_gain = gSum*( (g1/(g1+g2))*math.log2( g1/(g1+g2)) + (g2/(g1+g2))*math.log2( g2/(g1+g2))) + lSum*( (l1/(l1+l2))*math.log2( l1/(l1+l2)) + (l2/(l1+l2))*math.log2( l2/(l1+l2)))
  Info_gain = -Info_gain         
  Info_gain=Info_total-Info_gain
  gini = gSum*(1 - (g1/(g1+g2))**2 -(g2/(g1+g2))**2) +lSum*(1 - (l1/(l1+l2))**2 -(l2/(l1+l2))**2)

  print("Gini Index:"+str(gini))
  print("Information Gain:"+str(Info_gain))

else:
  #CODE FOR CATEGORICAL
  gini=0
  Info_gain=0

  categ = getUnique(x)
  categ = list(categ)

  for attr in categ:
    p1=0
    p2=0
    for i in range(2,row+1):
      val = sheet.cell(row=i,column=x).value
      if val==attr:
        if sheet.cell(row=i,column=y).value ==  Class[0]:
          p1=p1+1
        else:
          p2=p2+1

    sum=p1+p2
    a=0
    b=0
    if p1!=0:
      a=(p1/sum)*math.log2(p1/sum)
    if p2!=0:
      b=(p2/sum)*math.log2(p2/sum)
    Info_gain += ((sum)/n)*(a+b)
    gini += (sum/n)*(1- (p1/sum)**2 -(p2/sum)**2)
  
  Info_gain= -Info_gain
  Info_gain=Info_total-Info_gain
  print("Gini Index:"+str(gini))
  print("Information Gain:"+str(Info_gain))
