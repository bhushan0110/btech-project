import openpyxl
import math

wb= openpyxl.load_workbook("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment3.xlsx")
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

#calculating entropy of an array
def entropy(arr,l):
  one=0
  zero=0
  for i in range(0,l):
    if arr[i]==1:
      one = one+1
    else:
      zero = zero+1
  prob_one = one/l
  prob_zero= zero/l

  a=0
  b=0
  if prob_one!=0:
    a=prob_one*math.log(prob_one,2)
  if prob_zero!=0:
    b=prob_zero*math.log(prob_zero,2)
  return -(a+b)

#taking parent column
parent=[]
p_column=int(input("Enter parent column no: "))
for i in range(2,row+1):
  val = sheet_obj.cell(row=i,column=p_column)
  parent.append(val.value)

e_parent = entropy(parent,row-1)

#traversing each columns
for i in range(2,col+1):
  arr=[]
  if i!=p_column:
    for j in range(2,row+1):
      val=sheet_obj.cell(row=j,column=i)
      arr.append(val.value)
    #splitting column over parent column
    arr0=[]
    arr1=[]
    for k in range(0,row-1):
      if arr[k]==1:
        arr1.append(parent[k])
      else:
        arr0.append(parent[k])
    x=len(arr1)
    y=len(arr0)
    e_child1 = (entropy(arr1,x)*x)/(row-1)
    e_child2 = (entropy(arr0,y)*y)/(row-1)

    info_gain = e_parent-(e_child1+e_child2)
    val=sheet_obj.cell(row= row+2,column=i)
    val.value = info_gain

val = sheet_obj.cell(row=row+2,column=1)
val.value = "Information Gain"

wb.save("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment3_Ans.xlsx")