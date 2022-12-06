import openpyxl
import math 

def min_max(arr,minO,minN,maxO,maxN,l):
  arr1=[]
  for i in range(0,l):
    arr1.append((((arr[i]-minO)/(maxO-minO))*(maxN-minN)) + minN)
  # print(arr)
  return arr1

def z_score(arr,mn,sd,l):
  ans=[]
  for i in range(0,l):
    tmp=((arr[i]-mn)/sd)
    ans.append(tmp)
  return ans

def std_dev(arr,mn,l):
  sd=0
  for i in range(0,l):
    sd=sd+ (arr[i]-mn)*(arr[i]-mn)
  sd=math.sqrt(sd/l)
  return sd

def mean(arr,l):
  ans=0
  for i in range(0,l):
    ans=ans+arr[i]
  ans=ans/l
  return ans

def min_arr(arr,l):
  min=arr[0]
  for i in range(1,l):
    if arr[i]<min:
      min=arr[i]
  return min

def max_arr(arr,l):
  max=arr[0]
  for i in range(1,l):
    if arr[i]>max:
      max=arr[i]
  return max

wb_obj=openpyxl.load_workbook("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment2.xlsx")   
sheet_obj = wb_obj.active
row = sheet_obj.max_row
col = sheet_obj.max_column
minN=float(input("Enter new Min: "))
maxN=float(input("Enter new Max: "))
for i in range(2,col+1):
  arr=[]
  for j in range(2,row+1):
    val= sheet_obj.cell(row=j,column=i)
    arr.append(val.value)
  l=len(arr)
  minO=min_arr(arr,l)
  maxO=max_arr(arr,l)
  mn=mean(arr,l)
  sd=std_dev(arr,mn,l)
  MinMax=arr
  MinMax = min_max(MinMax,minO,minN,maxO,maxN,l)
  Z_Score = z_score(arr,mn,sd,l)

  row_mm=row+2
  row_zs=2*row+4
  for k in range(0,l):
    val=sheet_obj.cell(row=row_mm,column=1)
    val.value="Min_max"
    val=sheet_obj.cell(row=row_mm,column=i)
    val.value=MinMax[k]
    val=sheet_obj.cell(row=row_zs,column=1)
    val.value="Z_score"
    val=sheet_obj.cell(row=row_zs,column=i)
    val.value=Z_Score[k]
    row_mm=row_mm+1
    row_zs=row_zs+1
    
  
wb_obj.save("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment2_ans.xlsx")