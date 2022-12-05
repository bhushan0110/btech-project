import openpyxl
import math

wb=openpyxl.load_workbook("C:\\Users\\admin\\Desktop\\Bhushan\\DM\\assessment\\exp.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

wbook=openpyxl.load_workbook("C:\\Users\\admin\\Desktop\\Bhushan\\DM\\assessment\\ans.xlsx") 
sheet = wbook.active

def center(arr,l):
  c=0
  for i in range(0,l):
    c=c+arr[i]
  return c/l

def distance(x,y,xc,yc):
  d = (xc-x)*(xc-x) + (yc-y)*(yc-y)
  return math.sqrt(d)


X=[]
Y=[]
for i in range(2,row+1):
  valx = sheet_obj.cell(row=i,column=2)
  valy = sheet_obj.cell(row=i,column=3)
  X.append(valx.value)
  Y.append(valy.value)

length = len(X)

x_center=center(X,length)
y_center=center(Y,length)

################### CALCULATING MATRIX #####################
mat =[]
for i in range(0,length):
    row0=[]
    for j in range(0,length):
        row0.append(distance(X[i],Y[i],X[j],Y[j]))
    row0.append(distance(X[i],Y[i],x_center,y_center))
    mat.append(row0)
tmp=[]
for i in range(0,length):
    tmp.append(distance(x_center,y_center,X[i],Y[i]))
tmp.append(0)
mat.append(tmp)


############## WRITING DATA BACK TO THE FILE #############
x=len(mat)
i=0
for i in range(0,x):
    l=len(mat[i])
    for k in range(0,l):
        val = sheet.cell(row=(i+2),column=(k+2))
        val.value=round(mat[i][k],2)
        
array = ["P1","P2","P3","P4","Center"]
l= len(array)

for i in range(0,l):
    val = sheet.cell(row=1,column=i+2)
    val.value = array[i]
    val = sheet.cell(row=i+2,column=1)
    val.value = array[i]

################################################################
d = []
for i in range(0,length):
    d.append(distance(x_center,y_center,X[i],Y[i]))

### finding minimum distance from center
min=d[0]
ind=0
for i in range(1,length):
    if min>d[i]:
        min=d[i]
        ind=i

new_xc = X[ind]
new_yc = Y[ind]

final_d = []
for i in range(0,length):
    final_d.append(distance(new_xc,new_yc,X[i],Y[i]))
print(final_d)


new_x = []
new_y = []
for i in range(0,length):
    if i != ind:
        new_x.append(X[i])
        new_y.append(Y[i])

#################### CALCULATING NEW MATRIX #####################

matrix = []
new_length = len(new_x)
print(new_length)
for i in range(0,new_length):
    row0=[]
    for j in range(0,new_length):
        row0.append(distance(new_x[i],new_y[i],new_x[j],new_y[j]))
    row0.append(distance(new_x[i],new_y[i],new_xc,new_yc))
    matrix.append(row0)

tmp=[]
for i in range(0,new_length):
    tmp.append(distance(new_xc,new_xc,new_x[i],new_y[i]))
tmp.append(0)
matrix.append(tmp)

print(matrix)

###################### WRITING BACK NEW DATA ####################

i=0
for arr in matrix:
    ln=len(arr)
    for k in range(0,ln):
        val = sheet.cell(row=(i+12),column=(k+2))
        val.value=round(arr[k],2)
    i=i+1
# array = ["P1","P2","P3","Center"]
# l= len(array)

# for i in range(0,l):
#     val = sheet.cell(row=9,column=i+2)
#     val.value = array[i]
#     val = sheet.cell(row=i+9,column=1)
#     val.value = array[i]

wbook.save("C:\\Users\\admin\\Desktop\\Bhushan\\DM\\assessment\\ans.xlsx")