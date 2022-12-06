import openpyxl
import math
from itertools import combinations

wb=openpyxl.load_workbook("Experiment6.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

def read_data_in_dict(): 
  trans = []
  for i in range(2,row+1):
    x = sheet_obj.cell(row=i,column=1)
    trans.append(x.value)

  string_data=[]
  for i in range(2,row+1):
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        string_data.append(x.value)
  items = set(string_data)
  items = sorted(items)

  length = len(items)
  transactions=[]
  for i in range(2,row+1):
    s=[]
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        s.append(x.value)
    s=sorted(s)  
    req=[]
    k=0
    for j in range(0,length):
      if k<len(s):
        if items[j]==s[k]:
          req.append(1)
          k=k+1
        else:
          req.append(0)
      else:
        req.append(0)

    transactions.append(req)

  data={
      'items':items,
      'transactions': transactions
  }
  return data


data = read_data_in_dict()



def get_freq(s,items,transactions):
    freq=0
    for t in transactions:
        temp=1
        for item in s:
            temp*=t[items.index(item)]
        if temp==1:
            freq+=1  
    return freq



def frequent_itemsets(data,level,min_support):
    items = data['items']
    transactions = data['transactions']
    min_freq = math.ceil(min_support*len(transactions))
    sets = list(combinations(items,level))
    frequent_sets = []
    for s in sets:
        freq=get_freq(s,items,transactions)
        if freq>=min_freq:
            frequent_sets.append(s)
    return frequent_sets
        
        
x=len(data['items'])
min_sup=float(input("Enter Min Support: "))

val = sheet_obj.cell(row=row+2,column=1)
val.value = "Frequent Set:"

itr=row+3

print("Min Support: ",min_sup,"\n")
for i in range(1,x+1):
  arr=frequent_itemsets(data,i,min_sup)
  leng = len(arr)
  clmn=2
  if leng > 0:
   val=sheet_obj.cell(row=itr,column=1)
   val.value = "Level:"+str(i)
  for k in range(0,leng):
    val = sheet_obj.cell(row=itr,column=clmn)
    val.value=str(arr[k])
    clmn = clmn+1
  itr=itr+1

wb.save("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment6_Ans.xlsx")