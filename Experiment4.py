import openpyxl
import numpy as np

def unique(list1):
  x = np.unique(list1)
  print(x)
  return x


wb=openpyxl.load_workbook("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment4.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column
i=2
col1 = []
col2 = []
for i in range(2,row+1):
  val = sheet_obj.cell(row=i,column=1)
  col1.append(val.value)
  val = sheet_obj.cell(row=i,column=2)
  col2.append(val.value)

col1=unique(col1)
col2=unique(col2)

sum_t=[]
l=len(col1)
for i in range(0,l):
  tmp=0
  for j in range(2,row+1):
    val = sheet_obj.cell(row=j,column=1)
    val2= sheet_obj.cell(row=j,column=3)
    if col1[i]==val.value:
      tmp+=val2.value
  sum_t.append(tmp)


for i in range(0,l):
  for j in range(2,row+1):
    val = sheet_obj.cell(row=j,column=1)
    val2= sheet_obj.cell(row=j,column=3)
    if col1[i]==val.value:
      tweight = (val2.value/sum_t[i])*100
      x = sheet_obj.cell(row=j,column=4)
      x.value = tweight

sum_d=[]
l=len(col2)
for i in range(0,l):
  tmp=0
  for j in range(2,row+1):
    val = sheet_obj.cell(row=j,column=2)
    val2= sheet_obj.cell(row=j,column=3)
    if col2[i]==val.value:
      tmp+=val2.value
  sum_d.append(tmp)

for i in range(0,l):
  for j in range(2,row+1):
    val = sheet_obj.cell(row=j,column=2)
    val2= sheet_obj.cell(row=j,column=3)
    if col2[i]==val.value:
      dweight = (val2.value/sum_d[i])*100
      x = sheet_obj.cell(row=j,column=5)
      x.value = dweight

val=sheet_obj.cell(row=1,column=4)
val.value="T-weight"

val=sheet_obj.cell(row=1,column=5)
val.value="D-weight"

wb.save("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment4_Ans.xlsx") 