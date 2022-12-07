import openpyxl;

wb_obj = openpyxl.load_workbook("Bayes_classification.xlsx")
sheet = wb_obj.active
n = sheet.max_row
m=sheet.max_column

def getUnique(i):
    S=[]
    for j in range(2,n+1):
      S.append(sheet.cell(row=j,column=i).value)
    s=set(S)
    return s


b =input("Enter Class column no. : ")
b=int(b)
l=sheet.cell(row=1,column=b).value
print("Select attribute from Class column "+l+" :")
print(getUnique(b))

S = input("Enter value: ")


tot_mam=0
for i in range(1,n+1):
    if(sheet.cell(row=i,column=b).value==S):
        tot_mam+=1

list_mam=[]
print("\n\nEnter random case")
for i in range(2,m+1):
    if i!=b:
      k=sheet.cell(row=1,column=i).value
      print("Select the attribute req from below:")
      print(getUnique(i))
      a=input("Enter "+k+" : ")

      col_val=list(getUnique(i))
      cnt_arr=[]
      mammals=0
      for j in range(2,n+1):
        if(sheet.cell(row=j,column=i).value==a):
          if(sheet.cell(row=j,column=b).value==S):
            mammals=mammals+1
          b=int(b)
              
      list_mam.append(mammals/tot_mam)


prob_mam=1
for i in list_mam:
    prob_mam=prob_mam*i
prob_mam=prob_mam*(tot_mam/(n-1))


print("\nProbability for given conditon for "+l+" "+S+" is",prob_mam)
