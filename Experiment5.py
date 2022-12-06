import openpyxl
import math
def cal_median(array,l,r):     
  n = r+l
  median=0
  if n % 2 == 0:
    median=array[n//2]
  else:
    median1 = array[n//2]
    median2 = array[(n//2) + 1]
    
    median = (median1 + median2)/2
  return median

def five_no_summary(Arr):
  min=Arr[0]
  max=Arr[0]
  for x in Arr:
   if(x<min):
       min = x
  for x in Arr:
   if(x>max):
       max = x

  Arr.sort()
  n=len(Arr)
  #Median
  Med=cal_median(Arr,0,n-1)
  #Q1,Q3
  Q1=0
  Q3=0
  if(n%2==0):
    Q1=cal_median(Arr,0,(n//2)-1)
    Q3=cal_median(Arr,n//2,n-1)
  else:
    Q1=cal_median(Arr,0,(n//2))
    Q3=cal_median(Arr,(n)//2,n-1)

  #IQR
  IQR=Q3-Q1
  #Whiskers
  LW=Q1 - 1.5*IQR
  UW=Q3 + 1.5*IQR
    
  Ans=[]
  Ans.append(min)
  Ans.append(Q1)
  Ans.append(Med)
  Ans.append(Q3)
  Ans.append(max)
  Ans.append(IQR)
  Ans.append(LW)
  Ans.append(UW)
  return Ans

wb=openpyxl.load_workbook("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment5.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

i=1
  
# for i in range(1,col+1):

print(col)
for i in range(2,col+1):
  arr=[]
  for j in range(2,row+1):
    val = sheet_obj.cell(row = j, column = i)
    arr.append(val.value)
  ans=five_no_summary(arr)
  print(ans)
  k=row+2
  for x in ans:
    val=sheet_obj.cell(row=k,column=i)
    val.value=x
    k=k+1

arr1=["Min","Quatrile1","Median/Quatrile2","Quatrile3","Max","IQR","Lower Whisker","Upper Whisker"]
k=row+2
for x in arr1:
  val=sheet_obj.cell(row=k,column=1)
  val.value=x
  k=k+1


wb.save("C:\\Users\\Lenovo\\OneDrive\\Desktop\\DM\\Experiment5_Ans.xlsx")